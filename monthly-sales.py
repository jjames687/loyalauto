#!/usr/bin/env python

def print_intro():
    # Use a breakpoint in the code line below to debug your script.
    print("Welcome to Jesse's Spreadsheet Initializer")  # Press Ctrl+F8 to toggle the breakpoint.

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_intro()

#Here is where the Program starts
#Importing the datetime modules for Python
#import math module for Python
#import excel module for Python
from datetime import datetime, date
from datetime import timedelta
from datetime import date
from math import ceil
from openpyxl import Workbook


#Making a function to ask yes or no
def is_this_a_test():
    waiting = True
    while waiting == True:
        is_it = input("Do you want to change it?: Y or N?: ")
        if is_it == "Y" or is_it == "y":
            return True
        elif is_it == "N" or is_it == "n":
            return False


#This function queries and then returns the starting date
#Default to the first of the month. Thats how the spreadsheet is going to work
def starting_date():
    print("Pick a starting date: ")
    year = input("What year are we talking about here: ")
    print(" 1 Jan    2 Feb    3 Mar \n 4 Apr    5 May    6 Jun \n 7 Jul    8 Aug    9 Sep \n10 Oct   11 Nov   12 Dec")
    month = input("Which month? ")
    start_day = datetime(int(year), int(month), 1)
    print("You have selected: " + str(start_day)[0:10])
    return start_day

#Function to ask about monthly sales goal and ticket goal
def goals(b_list):
    print("There are " + str(len(b_list)) + " days in the month you have selected.")
    goal_monthly = input("What is your sales goal for this month: ")
    goal_daily = ceil(int(goal_monthly)/len(b_list))
    print("This means the daily goal will be $" + str(goal_daily))
    print("The default ticket goal is $500")
    goal_cars = ceil(goal_daily/500)
    print("This means the car goal will be " + str(goal_cars))
    return goal_daily, goal_cars

#This function makes a list of all the weekdays in the month chosen and then returns the list
def make_day_list(start):
    current_day = start
    day_list = []
    for x in range(1, 32):
        if current_day.weekday() <= 4 and (start.strftime("%m")) == (current_day.strftime("%m")):
            day_list.append(str(current_day))
        current_day += timedelta(days=1)
    for day in range(len(day_list)):
        day_list[day] = str(day_list[day])[0:10]
    return day_list

#This function will print the thing that I want to put into a spreadsheet anyway
def test_spreadsheet(a_list, sales_goal, cars_goal):
    a_list_of_days = a_list
    print("DATE      |SALES GOAL |SALES      |LESS |CARS GOAL |CARS      |TICKET GOAL |TICKET")
    multiplier = 1
    ticket_goal = 500
    pipe_string = "|           "
    for x in a_list:
        print(str(x[0:10]), end="")
        sales_string = str(sales_goal*multiplier)
        sales_string_length = len(sales_string)
        print(pipe_string[0:(11-sales_string_length)] + sales_string, end=" ")
        print(pipe_string + pipe_string[0:6], end="")
        cars_string = str(cars_goal*multiplier)
        cars_string_length = len(cars_string)
        print(pipe_string[0:(11-cars_string_length)] + cars_string, end="")
        print(pipe_string[0:11] + pipe_string[0:13-(len(str(ticket_goal)))] + str(ticket_goal) + pipe_string[0:7])
        multiplier +=1


# First pick the starting date
start = starting_date()

# Second make a list of dates starting with that date
the_list = make_day_list(start)

#We want to query the sales goal for the month and the ticket goal
#we have to do this after the list is made
the_daily_goal, the_car_goal = goals(the_list)

# Third print a mock spreadsheet
test_spreadsheet(the_list, the_daily_goal, the_car_goal)

#initialize spreadsheet
wb = Workbook()
dest_filename = input("What would you like to call the Spreadsheet?: ")
ws = wb.active
ws.title = "Sales Goals " + str(start)[0:10]

#column headers
headers = ["Date", "Sales Goal", "Sales", "Less", "Cars Goal", "Cars", "Ticket Goal", "Ticket"]
y = 0
for col in headers:
    cell_value = ws.cell(row=1, column=(y + 1), value=headers[y])
    y += 1

#value each column
y = 0
for row in the_list:
    cell_value = ws.cell(row=(2+y), column=1, value=the_list[y])
    y += 1
y = 0
for row in the_list:
    cell_value = ws.cell(row=(2+y), column=2, value=(the_daily_goal * (y+1)))
    y +=1
y = 0
for row in the_list:
    s_goal = ""
    s_goal = "B" + str(y + 2)
    s_act = ""
    s_act = "C" + str(y + 2)
    s_equation = ""
    s_equation = "=SUM(" + s_act + "-" + s_goal + ")"
    cell_value = ws.cell(row=(2 + y), column=4, value=s_equation)
    y += 1
y = 0
for row in the_list:
    cell_value = ws.cell(row=(2+y), column=5, value=(the_car_goal * (y+1)))
    y +=1
y = 0
for row in the_list:
    cell_value = ws.cell(row=(2+y), column=7, value="$500")
    y +=1

#savethatshit
wb.save(str(dest_filename) + '.xlsx')
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
